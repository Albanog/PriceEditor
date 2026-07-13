from __future__ import annotations

from openpyxl import load_workbook

from .models import Product


def import_excel(path: str) -> list[tuple[str, float]]:
    """Reads first two columns (Name, USD price) starting after header row."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 2:
            continue
        name, price = row[0], row[1]
        if name is None or price is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        try:
            price = float(price)
        except (TypeError, ValueError):
            continue
        rows.append((name, price))
    return rows


def merge_import(existing: list[Product], imported: list[tuple[str, float]]) -> list[Product]:
    """Update USD price on existing products matched by name (case-insensitive),
    keep their manual settings. Add new products for unmatched names.
    Existing products not present in the import are left untouched."""
    by_name = {p.name.strip().lower(): p for p in existing}
    for name, price in imported:
        key = name.strip().lower()
        if key in by_name:
            by_name[key].usd_price = price
        else:
            new_p = Product(name=name, usd_price=price)
            existing.append(new_p)
            by_name[key] = new_p
    return existing
